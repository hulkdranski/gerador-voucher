import customtkinter as ctk
import tkinter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from datetime import datetime, timedelta

#Definindo a aparencia padrão do sistema
ctk.set_appearance_mode('Dark')
ctk.set_default_color_theme('blue')

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.configuracao_layout()
        self.aparencia()
        self.todo_sistema()

    def configuracao_layout(self):
        self.title('Gerador de Voucher')
        self.geometry('500x300')
        self.resizable(False, False)

    def aparencia(self):
        #Aparencia do botao de temas
        ctk.CTkOptionMenu(self, values=['Dark', 'Light'], command=self.troca_aparencia).place(x=10, y=265)

    def troca_aparencia(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

    def todo_sistema(self):
        #primeiro frame
        frame = ctk.CTkFrame(self, width=500, height=50, corner_radius=0, bg_color='teal', fg_color='#30A6D9')
        frame.place(x=0, y=10)
        ctk.CTkLabel(frame, text='Gerador de Voucher', font=('Century Gothic bold', 20),
                             text_color='#fff').place(x=165, y=12)

        #frame_login
        self.frame_login = ctk.CTkFrame(self, width=500, height=200, corner_radius=0, bg_color='transparent', fg_color='transparent')
        self.frame_login.place(x=0, y=60)
        ctk.CTkLabel(self.frame_login, text='Login', font=('Century Gothic bold', 16)).place(x=230, y=5)
        login_entry = ctk.CTkEntry(self.frame_login)
        login_entry.place(x=180, y=35)
        ctk.CTkLabel(self.frame_login, text='Senha', font=('Century Gothic bold', 16)).place(x=230, y=75)
        senha_entry = ctk.CTkEntry(self.frame_login, show='*')
        senha_entry.place(x=180, y=105)
        validacao_login = ctk.CTkLabel(self.frame_login, text='', font=('Century Gothic bold', 16), text_color='red')
        validacao_login.place(x=160, y=135)

        ctk.CTkButton(self.frame_login, text='LOGIN', command=lambda:(login := login_entry.get(),
                                                                              senha := senha_entry.get(),
                                                                              self.logar(login, senha) if login == 'login_site' and senha == 'senha_site' else validacao_login.configure(text='Login ou senha incorretos')),
                      fg_color='#30A6D9', hover_color='#3071D9', font=('Century Gothic bold', 14)).place(x=180, y=170)


    def logar(self, login, senha):
        self.frame_login.destroy()
        text_var = ctk.BooleanVar()
        plan_var = ctk.BooleanVar()

        self.frame_gerar = ctk.CTkFrame(self, width=500, height=200, corner_radius=0, bg_color='transparent',
                                        fg_color='transparent')
        self.frame_gerar.place(x=0, y=10)
        ctk.CTkLabel(self.frame_gerar, text='Dias', font=('Century Gothic bold', 16)).place(x=85, y=5)
        dias_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        dias_entry.place(x=55, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Quantidade', font=('Century Gothic bold', 16)).place(x=210, y=5)
        quant_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        quant_entry.place(x=200, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Setor', font=('Century Gothic bold', 16)).place(x=375, y=5)
        setor_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        setor_entry.place(x=345, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Nome', font=('Century Gothic bold', 16)).place(x=345, y=75)
        nome_entry = ctk.CTkEntry(self.frame_gerar, width=150)
        nome_entry.place(x=295, y=105)
        ctk.CTkLabel(self.frame_gerar, text='Dispositivo', font=('Century Gothic bold', 16)).place(x=120, y=75)
        dispo_entry = ctk.CTkEntry(self.frame_gerar, width=210)
        dispo_entry.place(x=55, y=105)
        self.validacao_gerador = ctk.CTkLabel(self.frame_gerar, text='', font=('Century Gothic bold', 16), text_color='red')
        self.validacao_gerador.place(x=100, y=135)
        self.criar_plan = ctk.CTkCheckBox(self.frame_gerar, text='Salvar na planilha', onvalue=True, offvalue=False,
                                         variable=plan_var, border_width=1)
        self.criar_plan.place(x=55, y=168)
        self.criar_txt = ctk.CTkCheckBox(self.frame_gerar, text='Criar TXT', onvalue=True, offvalue=False, variable=text_var, border_width=1)
        self.criar_txt.place(x=190, y=168)
        botao_gerar = ctk.CTkButton(self.frame_gerar, text='GERAR', command=lambda:(dias := dias_entry.get(),
                                                                        quant := quant_entry.get(),
                                                                        dispo := dispo_entry.get(),
                                                                        setor := setor_entry.get(),
                                                                        nome := nome_entry.get(),
                                                                        self.gerar(login, senha, dias, quant, dispo, setor, nome, text_var, plan_var) if dias != '' and quant != '' and dispo != '' and nome != '' else self.validacao_gerador.configure(text='Todos campos são obrigatórios')), fg_color='#30A6D9',
                                         hover_color='#3071D9', font=('Century Gothic bold', 14), width=150)
        botao_gerar.place(x=300, y=165)

    def gerar(self, login, senha, dias, quant, dispo, setor, nome, text_var, plan_var):
        if int(quant) < 1 or int(quant) > 50:
            self.validacao_gerador.configure(text='Impossivel gerar essa quantidade de vouchers')
        elif int(dias) < 1:
            self.validacao_gerador.configure(text='Impossivel gerar essa quantidade de dias')
        else:
            options = webdriver.ChromeOptions()
            driver_service = Service(executable_path=r"chromedriver_path")
            driver = webdriver.Chrome(service=driver_service,
                                      options=options)
            driver.implicitly_wait(20)
            notas = ' - '.join([nome, setor, dispo])

            driver.get('login_page_unifi-network')

            # Clica em avançado
            driver.find_element('xpath', '/html/body/div/div[2]/button[3]').click()

            # Clica em ir para o site mesmo assim
            driver.find_element('xpath', '/html/body/div/div[3]/p[2]/a').click()

            # Coloca o usuario e senha
            driver.find_element('xpath',
                                '/html/body/div/ui-view/ui-view/div/div/div/div/div[3]/ui-view/div/form/div[1]/input').send_keys(
                login, Keys.ENTER)

            driver.find_element('xpath',
                                '/html/body/div/ui-view/ui-view/div/div/div/div/div[3]/ui-view/div/form/div[2]/input').send_keys(
                senha, Keys.ENTER)
            time.sleep(1)

            # Entra na área de vouncher
            driver.get('https://172.19.0.103:8443/manage/hotspot/site/default/vouchers/1/100')
            time.sleep(1)

            #pega um voucher "velho"
            nao_gerado = driver.find_element(By.XPATH,
                                             f'/html/body/div/ui-view/ui-view/div/div/div/ui-view/div/div[2]/div/table/tbody/tr[1]/td[2]').text

            # Clica em create voucher
            driver.find_element('xpath',
                                '/html/body/div[1]/ui-view/ui-view/div/div/div/ui-view/div/div[1]/div/div[1]/div/button[1]/span[2]').click()
            time.sleep(1)

            # coloca a quantidade de voucher
            driver.find_element('xpath',
                                '/html/body/div/div[4]/div[3]/div/div/div/div/form/div[2]/div/div/div[1]/div/div/input').send_keys(
                Keys.BACKSPACE, Keys.BACKSPACE, quant)

            # Coloca User-definet
            driver.find_element('xpath',
                                '/html/body/div[1]/div[4]/div[3]/div/div/div/div/form/div[2]/div/div/div[3]/div/div[1]/select[1]/option[7]').click()

            # colocar quantidade de dias
            driver.find_element('xpath',
                                '//input[@name="expire_number"]').send_keys(Keys.BACKSPACE, dias)

            # Coloca as notas
            driver.find_element('xpath',
                                '/html/body/div[1]/div[4]/div[3]/div/div/div/div/form/div[2]/div/div/div[7]/div/input').send_keys(
                notas)
            # Clica em save
            driver.find_element('xpath',
                                '/html/body/div/div[4]/div[3]/div/div/div/div/form/div[3]/div/div/div/button[2]').click()

            driver.refresh()

            numeros_voucher = []
            n = 0
            # Obter a data de hoje
            hoje = datetime.now()
            # Adicionar 75 dias à data de hoje
            data_futura = hoje + timedelta(days=int(dias))
            data_futura = data_futura.strftime("%d/%m/%Y")

            while n < int(quant):
                numeros_voucher.append(driver.find_element(By.XPATH,
                                                           f'/html/body/div/ui-view/ui-view/div/div/div/ui-view/div/div[2]/div/table/tbody/tr[{n + 1}]/td[2]').text)
                n += 1

            def copy_text(event):
                text = event.widget.cget("text")
                event.widget.focus_set()
                event.widget.clipboard_clear()
                event.widget.clipboard_append(text)

            def criar_label(frame, texto):
                label = ctk.CTkLabel(frame, text=texto)
                label.pack()
                label.configure(cursor="hand2")
                label.bind("<Button-1>", copy_text)

            def fechar_vouchers():
                self.frame_vouchers.destroy()
                self.logar(login, senha)

            def apresentar_vouchers():
                self.frame_vouchers = ctk.CTkScrollableFrame(self, width=500, height=200, corner_radius=0, bg_color='transparent',
                                                fg_color='transparent')
                self.frame_vouchers.place(x=0, y=10)
                for voucher in numeros_voucher:
                    criar_label(self.frame_vouchers, voucher)

                ctk.CTkButton(self.frame_vouchers, text='FECHAR', command=fechar_vouchers,
                                             fg_color='#30A6D9', hover_color='#3071D9', font=('Century Gothic bold', 14)).pack()

            if nao_gerado in numeros_voucher:
                self.validacao_gerador.configure(text='Não foi possível gerar o voucher')
            else:
                apresentar_vouchers()

            #Criar arquivo de texto
            if text_var.get():
                with open('numeros_voucher.txt', 'w') as arquivo:
                    for numero in numeros_voucher:
                        arquivo.write(numero + '\n')

            #Salvar na planilha
            def salvar_planilha():
                try:
                    # Carregar o arquivo Excel
                    wb = load_workbook(r'worksheet_to_save_voucher')

                    # Acessar a planilha
                    ws = wb['FUNCIONÁRIOS']

                    for cell in ws['B']:
                        if cell.value is None:
                            row = cell.row
                            for numero in numeros_voucher:
                                ws.cell(row=int(row), column=2, value=numero)
                                ws.cell(row=int(row), column=3, value=nome)
                                ws.cell(row=int(row), column=4, value=setor)
                                ws.cell(row=int(row), column=5, value=dispo)
                                ws.cell(row=int(row), column=6, value=data_futura)
                                row += 1
                            break

                    wb.save(r'worksheet_to_save_voucher')

                except PermissionError as e:
                    resposta = tkinter.messagebox.askyesno("Não foi possível colocar na planilha", f"{str(e)}\n\nNão foi possível colocar os dados do voucher na planilha. Possíveis causas: \n 1 - A planilha já está aberta \n 2 - Esse aplicativo não tem permissão para alterar a planilha\n\nTentar novamente?")

                    if resposta:
                        salvar_planilha()

            if plan_var.get():
                salvar_planilha()

if __name__ == "__main__":
    app = App()
    app.mainloop()